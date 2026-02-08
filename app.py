from pathlib import Path
from typing import Any, List, Optional, Literal
import importlib.util
from functools import lru_cache
import openpyxl
import sqlite3

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse
from pydantic import BaseModel, Field

# Pydantic v1/v2 совместимость
try:
    from pydantic import field_validator  # type: ignore
    _V2 = True
except Exception:
    from pydantic import validator as field_validator  # type: ignore
    _V2 = False

BASE_DIR = Path(__file__).resolve().parent
INDEX_HTML = BASE_DIR / "index.html"
INDEX_HTML_FALLBACK = BASE_DIR / "index-4.html"


def _load_rao_module():
    alt = BASE_DIR / "rao (2).py"
    if alt.exists():
        spec = importlib.util.spec_from_file_location("rao", alt)
        if spec is None or spec.loader is None:
            raise RuntimeError("Не удалось загрузить модуль rao")
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod
    import rao as mod  # type: ignore
    return mod


rao_mod = _load_rao_module()
run_calc_capture = rao_mod.run_calc_capture
parse_inn = rao_mod.parse_inn
get_org_name_by_inn = rao_mod.get_org_name_by_inn
fix_mojibake = rao_mod.fix_mojibake
load_licenses_by_inn = rao_mod.load_licenses_by_inn

RAO_DIR = Path(rao_mod.__file__).resolve().parent

def find_rkn_xlsx() -> Path:
    candidates = [
        BASE_DIR / "Таблица РКН slim.xlsx",
        BASE_DIR / "Таблица РКН.xlsx",
        BASE_DIR / "Таблица РКН (2).xlsx",
        BASE_DIR / "Таблица РКН очищенная.xlsx",
        RAO_DIR / "Таблица РКН.xlsx",
        RAO_DIR / "Таблица РКН (2).xlsx",
        RAO_DIR / "Таблица РКН slim.xlsx",
        RAO_DIR / "Таблица РКН очищенная.xlsx",
        Path.cwd() / "Таблица РКН.xlsx",
        Path.cwd() / "Таблица РКН (2).xlsx",
        Path.cwd() / "Таблица РКН slim.xlsx",
        Path.cwd() / "Таблица РКН очищенная.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(
        "Не найден файл 'Таблица РКН.xlsx'. "
        f"Пробовал: {[str(c) for c in candidates]}"
    )

def find_rkn_db() -> Optional[Path]:
    candidates = [
        BASE_DIR / "Таблица РКН.sqlite",
        BASE_DIR / "Таблица РКН.db",
        RAO_DIR / "Таблица РКН.sqlite",
        RAO_DIR / "Таблица РКН.db",
        Path.cwd() / "Таблица РКН.sqlite",
        Path.cwd() / "Таблица РКН.db",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None

def find_vars_xlsx() -> Path:
    candidates = [
        BASE_DIR / "Переменные из ставок.xlsx",
        RAO_DIR / "Переменные из ставок.xlsx",
        Path.cwd() / "Переменные из ставок.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(
        "Не найден файл 'Переменные из ставок.xlsx'. "
        f"Пробовал: {[str(c) for c in candidates]}"
    )


def _iter_rkn_rows_light(rkn_xlsx: Path):
    wb = openpyxl.load_workbook(rkn_xlsx, read_only=True, data_only=True)
    ws = wb.active
    header_raw = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    while header_raw and (header_raw[-1] is None or str(header_raw[-1]).strip() == ""):
        header_raw.pop()
    header = header_raw
    max_col = len(header)
    it = ws.iter_rows(min_row=2, max_col=max_col, values_only=True)
    return header, it


@lru_cache(maxsize=64)
def _licenses_light_cached(rkn_path: str, mtime: float, inn: str):
    rkn_xlsx = Path(rkn_path)
    header, it = _iter_rkn_rows_light(rkn_xlsx)
    idx = {h: i for i, h in enumerate(header)}

    def get(row, col):
        j = idx.get(col)
        if j is None:
            return None
        if j >= len(row):
            return None
        return row[j]

    by_license = {}
    for row in it:
        row_inn = str(get(row, "ns1:inn") or "").strip()
        if row_inn != inn:
            continue

        status = str(get(row, "ns1:status") or "").strip().lower()
        if status and status != "действующая":
            continue

        lic_id = str(get(row, "ns1:license_num") or "").strip()
        if not lic_id:
            continue

        org_name = str(get(row, "ns1:org_name") or "").strip()
        sreda = str(get(row, "ns1:sreda") or "").strip()
        pop_raw = get(row, "ns1:population")
        smi14 = str(get(row, "ns1:smi_name14") or "").strip()
        smi = str(get(row, "ns1:smi_name") or "").strip()

        lic = by_license.setdefault(lic_id, {
            "org_name": org_name,
            "media_raw": sreda,
            "pop_values": set(),
            "pop_notes": [],
            "smi_values": set(),
        })

        pop_int, pop_notes = rao_mod.parse_population(pop_raw)
        if pop_int is not None:
            lic["pop_values"].add(int(pop_int))
        if pop_notes:
            lic["pop_notes"].extend(pop_notes)

        if not lic["media_raw"] and sreda:
            lic["media_raw"] = sreda
        if smi14:
            lic["smi_values"].add(smi14)
        elif smi:
            lic["smi_values"].add(smi)

    items = []
    for lic_id, data in by_license.items():
        pop_total = None
        if data["pop_values"]:
            pop_total = int(sum(sorted(data["pop_values"])))
        media_raw = data.get("media_raw") or ""
        media_class = rao_mod.normalize_media(media_raw)
        smi_name = ""
        if data["smi_values"]:
            smi_name = sorted(data["smi_values"], key=lambda x: (len(x), x))[0]
        items.append({
            "license_id": lic_id,
            "media_raw": media_raw,
            "media_class": media_class,
            "population_total": pop_total,
            "population_notes": data.get("pop_notes", [])[:2],
            "channels_count": 0,
            "rkn_url": rao_mod.build_rkn_url(lic_id),
            "org_name": data.get("org_name", ""),
            "smi_name": smi_name,
        })
    return items


def load_licenses_light(rkn_xlsx: Path, inn: str):
    return _licenses_light_cached(str(rkn_xlsx), rkn_xlsx.stat().st_mtime, inn)


_RKN_INDEX = None
_RKN_INDEX_MTIME = None
_RKN_INDEX_SRC = None


def _build_rkn_index_from_sqlite(db_path: Path):
    idx = {}
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    for row in cur.execute(
        "SELECT inn, org_name, license_num, sreda, population, status, smi_name14, smi_name FROM rkn"
    ):
        inn, org_name, lic_id, sreda, pop_raw, status, smi14, smi = row
        inn = (inn or "").strip()
        if not inn:
            continue
        st = (status or "").strip().lower()
        if st and st != "действующая":
            continue
        lic_id = (lic_id or "").strip()
        if not lic_id:
            continue
        by_lic = idx.setdefault(inn, {})
        lic = by_lic.setdefault(
            lic_id,
            {
                "org_name": (org_name or "").strip(),
                "media_raw": (sreda or "").strip(),
                "pop_values": set(),
                "pop_notes": [],
                "smi_values": set(),
            },
        )
        pop_int, pop_notes = rao_mod.parse_population(pop_raw)
        if pop_int is not None:
            lic["pop_values"].add(int(pop_int))
        if pop_notes:
            lic["pop_notes"].extend(pop_notes)
        if smi14:
            lic["smi_values"].add(str(smi14).strip())
        elif smi:
            lic["smi_values"].add(str(smi).strip())
    conn.close()
    return idx


def _build_rkn_index_from_xlsx(rkn_xlsx: Path):
    header, it = _iter_rkn_rows_light(rkn_xlsx)
    idx = {h: i for i, h in enumerate(header)}

    def get(row, col):
        j = idx.get(col)
        if j is None:
            return None
        if j >= len(row):
            return None
        return row[j]

    out = {}
    for row in it:
        inn = str(get(row, "ns1:inn") or "").strip()
        if not inn:
            continue
        status = str(get(row, "ns1:status") or "").strip().lower()
        if status and status != "действующая":
            continue
        lic_id = str(get(row, "ns1:license_num") or "").strip()
        if not lic_id:
            continue
        org_name = str(get(row, "ns1:org_name") or "").strip()
        sreda = str(get(row, "ns1:sreda") or "").strip()
        pop_raw = get(row, "ns1:population")
        smi14 = str(get(row, "ns1:smi_name14") or "").strip()
        smi = str(get(row, "ns1:smi_name") or "").strip()

        by_lic = out.setdefault(inn, {})
        lic = by_lic.setdefault(
            lic_id,
            {
                "org_name": org_name,
                "media_raw": sreda,
                "pop_values": set(),
                "pop_notes": [],
                "smi_values": set(),
            },
        )
        pop_int, pop_notes = rao_mod.parse_population(pop_raw)
        if pop_int is not None:
            lic["pop_values"].add(int(pop_int))
        if pop_notes:
            lic["pop_notes"].extend(pop_notes)
        if smi14:
            lic["smi_values"].add(smi14)
        elif smi:
            lic["smi_values"].add(smi)
    return out


def get_licenses_index():
    global _RKN_INDEX, _RKN_INDEX_MTIME, _RKN_INDEX_SRC
    db = find_rkn_db()
    if db:
        mtime = db.stat().st_mtime
        if (
            _RKN_INDEX is None
            or _RKN_INDEX_MTIME != mtime
            or _RKN_INDEX_SRC != str(db)
        ):
            _RKN_INDEX = _build_rkn_index_from_sqlite(db)
            _RKN_INDEX_MTIME = mtime
            _RKN_INDEX_SRC = str(db)
        return _RKN_INDEX

    rkn_xlsx = find_rkn_xlsx()
    mtime = rkn_xlsx.stat().st_mtime
    if (
        _RKN_INDEX is None
        or _RKN_INDEX_MTIME != mtime
        or _RKN_INDEX_SRC != str(rkn_xlsx)
    ):
        _RKN_INDEX = _build_rkn_index_from_xlsx(rkn_xlsx)
        _RKN_INDEX_MTIME = mtime
        _RKN_INDEX_SRC = str(rkn_xlsx)
    return _RKN_INDEX

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://ksenonya.github.io",
        "https://rnk-wof7.onrender.com",
        "http://localhost",
        "http://127.0.0.1:8000",
    ],
    allow_origin_regex=r"https://.*\\.github\\.io",
    allow_methods=["*"],
    allow_headers=["*"],
)

INN_MAP: dict[str, str] = {}

@app.on_event("startup")
def load_inn_map():
    global INN_MAP
    csv_path = BASE_DIR / "inn_name.csv"
    if not csv_path.exists():
        print("⚠️ inn_name.csv not found рядом с app.py")
        return

    import csv
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f)

        # 🔥 проверка, что колонки совпали
        if not r.fieldnames:
            raise RuntimeError("inn_name.csv пустой или без заголовков")
        need_cols = {"ns1:inn", "ns1:org_name_short"}
        missing = need_cols - set(r.fieldnames)
        if missing:
            raise RuntimeError(f"inn_name.csv: не найдены колонки {sorted(missing)}; есть {r.fieldnames}")

        mp = {}
        for row in r:
            inn = (row.get("ns1:inn") or "").strip()
            name = (row.get("ns1:org_name_short") or "").strip()
            if inn and name and inn not in mp:
                mp[inn] = name

        INN_MAP = mp
        print(f"✅ inn_name.csv loaded: {len(INN_MAP)}")



DASH_TOKENS = {"", "-", "—", "–", "нет"}


def _is_dash(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip().lower() in DASH_TOKENS
    return False


def _to_none_or_str(v: Any) -> Optional[str]:
    if _is_dash(v):
        return None
    return str(v).strip()


def _to_none_or_int(v: Any) -> Optional[int]:
    if _is_dash(v):
        return None
    if isinstance(v, int):
        return v
    s = str(v).strip().replace(" ", "")
    if not s:
        return None
    return int(s)


def _clean_license_ids(v: Any) -> Optional[List[str]]:
    if v is None:
        return None
    if isinstance(v, str):
        parts = [p.strip() for p in v.split(",")]
        items = [p for p in parts if p]
    else:
        try:
            items = [str(x).strip() for x in list(v)]
        except Exception:
            items = []
    items = [x for x in items if x and not _is_dash(x)]
    return items or None


def _clean_population_by_license(v: Any) -> Optional[dict]:
    if v is None:
        return None
    if not isinstance(v, dict):
        return None
    out: dict[str, int] = {}
    for k, val in v.items():
        key = _to_none_or_str(k)
        if not key:
            continue
        n = _to_none_or_int(val)
        if n is None:
            continue
        out[str(key)] = int(n)
    return out or None


def _to_none_or_float(v: Any) -> Optional[float]:
    if _is_dash(v):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(",", ".")
    if not s:
        return None
    return float(s)


class CalcRequest(BaseModel):
    inn: str = Field(..., description="ИНН 10 или 12 цифр")

    annual_revenue: Optional[float] = Field(None, ge=0)
    revenue_q: Optional[float] = Field(None, ge=0)
    expenses_q: Optional[float] = Field(None, ge=0)

    internet_resources: int = Field(0, ge=0, le=1000)
    contract_media: Literal["auto", "cable", "air", "both"] = "auto"

    new_user: bool = False
    assoc_member: bool = False

    only_license: Optional[str] = None
    license_ids: Optional[List[str]] = None
    population_override: Optional[int] = Field(None, ge=0, le=2_000_000_000)
    population_by_license: Optional[dict[str, int]] = None
    subscriber_total: Optional[int] = Field(None, ge=0, le=2_000_000_000)

    if _V2:
        @field_validator("inn", mode="before")
        @classmethod
        def _v_inn(cls, v: Any) -> str:
            s = _to_none_or_str(v)
            if not s:
                raise ValueError("ИНН обязателен")
            s = s.replace(" ", "")
            if not s.isdigit() or len(s) not in (10, 12):
                raise ValueError("ИНН должен состоять из 10 или 12 цифр")
            return s

        @field_validator("annual_revenue", "revenue_q", "expenses_q", mode="before")
        @classmethod
        def _v_floats(cls, v: Any) -> Optional[float]:
            return _to_none_or_float(v)

        @field_validator("only_license", mode="before")
        @classmethod
        def _v_only_license(cls, v: Any) -> Optional[str]:
            return _to_none_or_str(v)

        @field_validator("license_ids", mode="before")
        @classmethod
        def _v_license_ids(cls, v: Any) -> Optional[List[str]]:
            return _clean_license_ids(v)

        @field_validator("population_override", mode="before")
        @classmethod
        def _v_pop(cls, v: Any) -> Optional[int]:
            return _to_none_or_int(v)

        @field_validator("population_by_license", mode="before")
        @classmethod
        def _v_pop_by_license(cls, v: Any) -> Optional[dict]:
            return _clean_population_by_license(v)

    else:
        @field_validator("inn", pre=True)
        def _v1_inn(cls, v: Any) -> str:
            s = _to_none_or_str(v)
            if not s:
                raise ValueError("ИНН обязателен")
            s = s.replace(" ", "")
            if not s.isdigit() or len(s) not in (10, 12):
                raise ValueError("ИНН должен состоять из 10 или 12 цифр")
            return s

        @field_validator("annual_revenue", "revenue_q", "expenses_q", pre=True)
        def _v1_floats(cls, v: Any) -> Optional[float]:
            return _to_none_or_float(v)

        @field_validator("only_license", pre=True)
        def _v1_only_license(cls, v: Any) -> Optional[str]:
            return _to_none_or_str(v)

        @field_validator("license_ids", pre=True)
        def _v1_license_ids(cls, v: Any) -> Optional[List[str]]:
            return _clean_license_ids(v)

        @field_validator("population_override", pre=True)
        def _v1_pop(cls, v: Any) -> Optional[int]:
            return _to_none_or_int(v)

        @field_validator("population_by_license", pre=True)
        def _v1_pop_by_license(cls, v: Any) -> Optional[dict]:
            return _clean_population_by_license(v)

@app.get("/", response_class=HTMLResponse)
def home():
    if not INDEX_HTML.exists() and INDEX_HTML_FALLBACK.exists():
        return HTMLResponse(INDEX_HTML_FALLBACK.read_text(encoding="utf-8"))
    if not INDEX_HTML.exists():
        return HTMLResponse(
            "<h1>index.html не найден</h1><p>Положи index.html рядом с app.py</p>",
            status_code=500,
        )
    return HTMLResponse(INDEX_HTML.read_text(encoding="utf-8"))




@app.get("/api/inninfo")
def api_inninfo(inn: str):
    try:
        inn_clean = parse_inn(inn)
        org_name = INN_MAP.get(inn_clean)
        if org_name:
            org_name = fix_mojibake(org_name)
            return {"ok": True, "inn": inn_clean, "org_name": org_name}
    except Exception as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})

    try:
        rkn_xlsx = find_rkn_xlsx()
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})

    org_name = get_org_name_by_inn(rkn_xlsx, inn_clean)
    org_name = fix_mojibake(org_name)

    if not org_name:
        return JSONResponse(status_code=404, content={"ok": False, "org_name": ""})

    return {"ok": True, "inn": inn_clean, "org_name": org_name}


@app.get("/api/licenses")
def api_licenses(inn: str):
    try:
        inn_clean = parse_inn(inn)
    except Exception as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})

    try:
        rkn_xlsx = find_rkn_xlsx()
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})

    # Быстрая загрузка из индекса (sqlite/память)
    try:
        index = get_licenses_index()
        by_lic = index.get(inn_clean, {})
        items = []
        for lic_id, data in by_lic.items():
            pop_total = None
            if data["pop_values"]:
                pop_total = int(sum(sorted(data["pop_values"])))
            media_raw = data.get("media_raw") or ""
            media_class = rao_mod.normalize_media(media_raw)
            smi_name = ""
            if data.get("smi_values"):
                smi_name = sorted(data["smi_values"], key=lambda x: (len(x), x))[0]
            items.append({
                "license_id": lic_id,
                "media_raw": media_raw,
                "media_class": media_class,
                "population_total": pop_total,
                "population_notes": data.get("pop_notes", [])[:2],
                "channels_count": 0,
                "rkn_url": rao_mod.build_rkn_url(lic_id),
                "org_name": data.get("org_name", ""),
                "smi_name": smi_name,
            })
        return {"ok": True, "inn": inn_clean, "licenses": items, "notes": []}
    except Exception as e:
        # fallback на «лёгкую» загрузку без индекса
        try:
            items = load_licenses_light(rkn_xlsx, inn_clean)
            return {"ok": True, "inn": inn_clean, "licenses": items, "notes": []}
        except Exception as e2:
            return JSONResponse(status_code=500, content={"ok": False, "error": f"Ошибка загрузки лицензий: {e} / fallback: {e2}"})


@app.post("/api/calc")
@app.post("/api/calc/")
def api_calc(req: CalcRequest):
    argv: List[str] = ["--inn", req.inn.strip(), "--non_interactive"]

    if req.revenue_q is not None:
        argv += ["--revenue_q", str(req.revenue_q)]
    if req.annual_revenue is not None:
        argv += ["--annual_revenue", str(req.annual_revenue)]
    if req.expenses_q is not None:
        argv += ["--expenses_q", str(req.expenses_q)]

    argv += ["--internet_resources", str(req.internet_resources)]
    argv += ["--contract_quarter", "1"]
    argv += ["--contract_media", str(req.contract_media)]

    if req.new_user:
        argv.append("--new_user")
    if req.assoc_member:
        argv.append("--assoc_member")

    if req.only_license:
        argv += ["--only_license", req.only_license.strip()]
    if req.license_ids:
        for lic in req.license_ids:
            if lic:
                argv += ["--licenses", str(lic).strip()]
    if req.population_override is not None:
        argv += ["--population_override", str(int(req.population_override))]
    if req.population_by_license:
        for lic, pop in req.population_by_license.items():
            argv += ["--population_by_license", f"{lic}={int(pop)}"]
    if req.subscriber_total is not None:
        argv += ["--subscriber_total", str(int(req.subscriber_total))]


    code, out = run_calc_capture(argv)
    out = (out or "").strip()

    if code == 0:
        return JSONResponse(status_code=200, content={"ok": True, "text": out})
    return JSONResponse(status_code=400, content={"ok": False, "error": out or "Ошибка расчёта"})


@app.get("/api/version")
def api_version():
    return {"app": "app_fix", "rao": "rao_fix2"}
