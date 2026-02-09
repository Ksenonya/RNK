# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import io
import math
import re
from contextlib import redirect_stdout
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote
import unicodedata

from functools import lru_cache

import openpyxl
import pandas as pd


# --------------------------- helpers: progress ---------------------------

class Progress:
    def __init__(self, enabled: bool = True) -> None:
        self.enabled = enabled
        self.total = 10
        self.step = 0

    def tick(self, msg: str) -> None:
        if not self.enabled:
            return
        self.step = min(self.total, self.step + 1)
        filled = int((self.step / self.total) * 10)
        bar = "■" * filled + "□" * (10 - filled)
        pct = int((self.step / self.total) * 100)
        print(f"Прогресс: [{bar}] {pct}% — {msg}")


# --------------------------- parsing ---------------------------

def parse_inn(raw: str) -> str:
    s = re.sub(r"\D+", "", raw or "")
    if len(s) not in (10, 12):
        raise ValueError("ИНН должен состоять из 10 или 12 цифр.")
    return s


def parse_int_like(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        if math.isnan(v):
            return None
        return int(round(v))
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("\u00a0", " ")
    s = s.replace(" ", "").replace(",", ".")
    if not re.match(r"^-?\d+(\.\d+)?$", s):
        return None
    return int(round(float(s)))


def parse_population(v: Any) -> Tuple[Optional[int], List[str]]:
    notes: List[str] = []
    if v is None:
        return None, notes

    if isinstance(v, int):
        return int(v), notes
    if isinstance(v, float):
        if math.isnan(v):
            return None, notes
        if v < 10000 and abs(v - round(v)) > 1e-9:
            notes.append("Население было дробным числом без единиц; применена эвристика «тыс.» (×1000).")
            return int(round(v * 1000)), notes
        return int(round(v)), notes

    s = str(v).strip().lower().replace("\u00a0", " ")
    if not s:
        return None, notes

    mult = 1
    if "млн" in s:
        mult = 1_000_000
    elif "тыс" in s:
        mult = 1_000

    num = re.sub(r"[^0-9,.\- ]+", "", s)
    num = num.replace(" ", "").replace(",", ".")
    if not num:
        return None, notes

    try:
        f = float(num)
    except ValueError:
        # попытка обработать диапазоны/смешанные значения (например "10-20 тыс")
        nums = re.findall(r"\d+(?:[.,]\d+)?", s)
        if not nums:
            return None, notes
        vals = []
        for x in nums:
            try:
                vals.append(float(x.replace(",", ".")))
            except ValueError:
                continue
        if not vals:
            return None, notes
        f = max(vals)
        notes.append("Население было диапазоном/смешанным значением; взято максимальное.")

    if mult == 1 and f < 10000 and abs(f - round(f)) > 1e-9:
        notes.append("Население было дробным числом без единиц; применена эвристика «тыс.» (×1000).")
        mult = 1_000

    return int(round(f * mult)), notes


def parse_license_list(items: Optional[List[str]], only_license: Optional[str]) -> List[str]:
    out: List[str] = []
    if items:
        for raw in items:
            if raw is None:
                continue
            parts = re.split(r"[,\n]+", str(raw))
            for p in parts:
                s = p.strip()
                if s:
                    out.append(s)
    if only_license:
        s = str(only_license).strip()
        if s:
            out.append(s)

    # unique, preserve order
    seen = set()
    uniq: List[str] = []
    for x in out:
        if x in seen:
            continue
        seen.add(x)
        uniq.append(x)
    return uniq


def parse_population_by_license(items: Optional[List[str]]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    if not items:
        return out
    for raw in items:
        if raw is None:
            continue
        parts = re.split(r"[,\n]+", str(raw))
        for p in parts:
            s = p.strip()
            if not s:
                continue
            if "=" not in s:
                continue
            lic, val = s.split("=", 1)
            lic = lic.strip()
            if not lic:
                continue
            pop, _ = parse_population(val)
            if pop is None:
                continue
            out[lic] = int(pop)
    return out


def parse_hours_week(brcst_time: Any, smi_name: Any) -> Tuple[Optional[float], List[str]]:
    notes: List[str] = []

    if brcst_time is not None and str(brcst_time).strip() != "":
        s = str(brcst_time).strip().lower()
        if "кругл" in s:
            return 168.0, notes
        n = parse_int_like(s)
        if n is not None:
            return float(n), notes

    if smi_name:
        m = re.search(r"\((\d{1,3})\)", str(smi_name))
        if m:
            return float(int(m.group(1))), notes

    return None, notes


def normalize_media(sreda: Any) -> str:
    s = (str(sreda or "")).lower()
    has_air = ("эфир" in s) or ("назем" in s)
    has_cable = ("кабель" in s)
    has_univ = ("универс" in s)
    if has_univ:
        return "Одновременно в эфире и по кабелю"
    if has_air and has_cable:
        return "Одновременно в эфире и по кабелю"
    return "В эфире или по кабелю"


def clean_channel_name(name: Any) -> str:
    s = str(name or "").strip()
    if not s:
        return ""
    s = re.sub(r"\s*\(\d{1,3}\)\s*$", "", s).strip()
    return s


# --------------------------- models ---------------------------

@dataclass
class TopicShare:
    topic_raw: str
    share_pct: Optional[float]
    rate_pct: float
    note: Optional[str] = None


@dataclass
class Channel:
    name: str
    hours_week: Optional[float]
    topics: List[TopicShare] = field(default_factory=list)

    def avg_rate(self) -> Tuple[float, List[str]]:
        notes: List[str] = []
        if not self.topics:
            notes.append("Тематики не найдены; ставка по умолчанию 2,5%.")
            return 2.5, notes

        shares = [t for t in self.topics if t.share_pct is not None]
        if shares:
            for t in shares:
                if t.share_pct > 50:
                    notes.append(f"Преобладающая тематика >50%: «{t.topic_raw}» ({t.share_pct}%).")
                    return t.rate_pct, notes

            total = sum(t.share_pct for t in shares)
            if total > 0:
                wavg = sum(t.share_pct * t.rate_pct for t in shares) / total
                notes.append("Ставка телеканала рассчитана как взвешенное среднее по долям тематик.")
                return wavg, notes

        avg = sum(t.rate_pct for t in self.topics) / len(self.topics)
        notes.append("Доли тематик отсутствуют/неполные; ставка телеканала рассчитана как простое среднее.")
        return avg, notes


@dataclass
class License:
    license_id: str
    org_name: str
    inn: str
    media_raw: str
    media_class: str
    population_total: Optional[int]
    population_notes: List[str] = field(default_factory=list)
    rkn_url: str = ""
    channels: List[Channel] = field(default_factory=list)

    def total_hours(self) -> float:
        hrs = [c.hours_week for c in self.channels if c.hours_week is not None]
        if hrs:
            return float(sum(hrs))
        return 168.0


# --------------------------- topic -> rate ---------------------------

DEFAULT_TOPIC_RATE = 2.5


TOPIC_MAP_COL_CAT = "Категория тематики использования произведений по Приложению 1"
TOPIC_MAP_COL_TOPIC = "Формулировка тематики вещания в лицензии пользователя"


def normalize_topic(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("ё", "е")
    # разные тире -> дефис
    s = s.replace("—", "-").replace("–", "-")
    # unicode normalize
    s = unicodedata.normalize("NFKC", s)
    # убрать скобки/кавычки/пунктуацию в пробел
    s = re.sub(r"[\"'«»“”]", " ", s)
    s = re.sub(r"[()\[\]{}]", " ", s)
    s = re.sub(r"[^a-zа-я0-9\- ]+", " ", s)
    # схлопнуть пробелы
    s = re.sub(r"\s+", " ", s).strip()
    return s


# Дополнительные тематики из перечня (скриншоты РАО).
EXTRA_TOPICS_BY_CATEGORY: Dict[str, List[str]] = {
    "I": [
        "аналитическое",
        "информационно-новостное",
        "вещание в сфере бизнеса",
        "военное",
        "военно-патриотическое",
        "духовно-просветительское",
        "интервью",
        "информационно-аналитическое",
        "информационно-деловое",
        "информационное",
        "информационно-публицистическое",
        "информационно-справочное",
        "информационно-экономическое",
        "информационные выпуски",
        "обзоры новостей",
        "информационные рубрики",
        "информационные передачи (программы)",
        "комментарии",
        "новостное (любой тематики)",
        "новостное",
        "общественно-информационное",
        "общественно-политическое",
        "общественно-социальное",
        "оперативная информация для водителей",
        "официальная хроника и публицистика",
        "патриотическое",
        "политическое",
        "правовое",
        "производственно-экономическое",
        "публицистическо-аналитическое",
        "публицистическое",
        "религиозное",
        "религиозно-просветительское",
        "сельскохозяйственное",
        "сообщения и объявления",
        "официальные",
        "социальное",
        "социально-значимые передачи",
        "социально-публицистическое",
        "социально-экономическое",
        "специализированное информационное",
        "справочное",
        "экологическое",
        "экономическое",
    ],
    "II": [
        "детское",
        "для детей",
        "для подростков",
        "для школьников",
        "информационно-культурное",
        "информационно-познавательное",
        "информационно-спортивное",
        "культурно-публицистическое",
        "литературно-публицистическое",
        "научное",
        "научно-образовательное",
        "научно-познавательное",
        "научно-популярное",
        "образовательное",
        "образовательные передачи (программы)",
        "подростковое",
        "познавательное",
        "пропаганда здорового образа жизни",
        "просветительское",
        "просветительско-образовательное",
        "разговорное",
        "разговорные передачи",
        "разговорные передачи или программы",
        "семейное",
        "спортивное",
        "спортивно-оздоровительное",
        "спортивные передачи (программы)",
        "туризм",
        "уроки",
        "учебно-познавательное",
        "учебно-просветительское",
        "художественно-политическое",
        "юношеское",
    ],
    "III": [
        "досуг",
        "искусство",
        "культурное",
        "культурно-просветительское",
        "культурно-развлекательное",
        "литературно-поэтическое",
        "литературно-художественное",
        "литературно-художественные программы",
        "народное творчество",
        "отдых",
        "передачи об искусстве",
        "прогноз погоды",
        "социально-культурное",
        "спортивно-развлекательное",
        "тематические передачи (программы)",
        "художественное",
        "художественно-публицистическое",
        "художественные и (или) документальные кино- и телефильмы",
        "художественные передачи (программы)",
        "документальные кино- и телефильмы",
    ],
    "IV": [
        "викторины",
        "игры",
        "информационно-развлекательное",
        "информационно-рекламное",
        "конкурсы",
        "литературно-драматическое",
        "молодежное",
        "молодежные и развлекательные передачи (программы)",
        "передачи развлекательного характера",
        "поздравительно-развлекательное",
        "поздравления",
        "познавательно-развлекательное",
        "развлекательно-воспитательное",
        "развлекательное",
        "развлекательно-игровое",
        "развлекательно-информационное",
        "развлекательно-познавательное",
        "развлекательные передачи (программы)",
        "развлечения",
        "рекламно-информационное",
        "ток-шоу",
        "ток шоу",
        "шоу",
        "шоу-программы",
        "юмористические передачи (программы)",
        "юмористическое",
    ],
    "V": [
        "концерты, в т.ч. по заявкам",
        "концерты",
        "литературно-музыкальное",
        "музыка",
        "музыкальное",
        "музыкально-информационное",
        "информационно-музыкальное",
        "музыкально-информационные передачи (программы)",
        "развлекательного характера (программы)",
        "музыкально-информационно-развлекательное",
        "музыкально-поздравительное",
        "музыкально-развлекательное",
        "музыкально-развлекательные передачи (программы)",
        "музыкально-тематические передачи",
        "музыкальные и развлекательные передачи",
        "музыкальные конкурсы",
        "музыкальные новости",
        "музыкальные передачи (программа)",
        "передачи (программы) о музыке",
        "песни",
        "популярная музыка",
        "развлекательно-музыкальное",
        "реклама",
        "рекламное",
        "рекламно-развлекательное",
        "рекламные ролики",
        "рекламные сообщения и материалы",
        "сюжеты на правах рекламы",
        "трансляция музыкальных передач (программ), концертов, фестивалей, праздников и других передач музыкального содержания",
        "трансляция музыкальных передач",
        "передачи музыкального содержания",
    ],
}


def _build_extra_topics_df() -> pd.DataFrame:
    rows: List[Dict[str, str]] = []
    for cat, topics in EXTRA_TOPICS_BY_CATEGORY.items():
        for t in topics:
            t = str(t or "").strip()
            if not t:
                continue
            rows.append({
                TOPIC_MAP_COL_TOPIC: t,
                TOPIC_MAP_COL_CAT: cat,
            })
    if not rows:
        return pd.DataFrame(columns=[TOPIC_MAP_COL_TOPIC, TOPIC_MAP_COL_CAT])
    return pd.DataFrame(rows)


EXTRA_TOPICS_DF = _build_extra_topics_df()

def build_category_rate_map(vars_xlsx: Path) -> Dict[str, float]:
    df = pd.read_excel(vars_xlsx, sheet_name="Категории и ставки")
    out: Dict[str, float] = {}
    for _, r in df.iterrows():
        cat = str(r.get("Категория использования произведений (по Приложению 1)", "")).strip()
        rate = r.get("Ставка авторского вознаграждения, процентов от дохода или расходов")
        if cat and pd.notna(rate):
            out[cat] = float(rate)
    return out


def topic_to_rate(topic: str, category_rate: Dict[str, float], mapping_df: Optional[pd.DataFrame]) -> Tuple[float, List[str]]:
    notes: List[str] = []

    t_raw = (topic or "").strip()
    tl = normalize_topic(t_raw)

    # ---------------- 1) Пытаемся сопоставить по таблице (не только ==, но и contains) ----------------
    if mapping_df is not None and not mapping_df.empty:
        df = mapping_df.copy()
    else:
        df = pd.DataFrame(columns=[TOPIC_MAP_COL_TOPIC, TOPIC_MAP_COL_CAT])

    if not EXTRA_TOPICS_DF.empty:
        df = pd.concat([df, EXTRA_TOPICS_DF], ignore_index=True)

    if TOPIC_MAP_COL_CAT in df.columns and TOPIC_MAP_COL_TOPIC in df.columns:
        df = df[[TOPIC_MAP_COL_TOPIC, TOPIC_MAP_COL_CAT]].dropna()
        if not df.empty:
            df = df.copy()
            df["_norm"] = df[TOPIC_MAP_COL_TOPIC].astype(str).map(normalize_topic)

            # 1) точное совпадение
            m = df[df["_norm"] == tl]
            if not m.empty:
                cat = str(m.iloc[0][TOPIC_MAP_COL_CAT]).strip()
                rate = category_rate.get(cat)
                if rate is not None:
                    notes.append(f"Тематика сопоставлена по таблице «Тематики по категориям» (точно): категория {cat}.")
                    return float(rate), notes

            # 2) “вхождение” в обе стороны (часто формулировки длиннее/короче)
            # берем самый “длинный” матч (обычно он точнее)
            # 2) “вхождение” в обе стороны (часто формулировки длиннее/короче)
            # 2) “вхождение” в обе стороны (длиннее/короче)
            # mask1: в лицензии tl содержит формулировку из таблицы
            # mask2: формулировка из таблицы содержит tl
            if tl:  # защита от пустой строки
                # сначала ищем формулировки, которые входят в тематику (точнее)
                mask_in_tl = df["_norm"].apply(lambda x: bool(x) and (x in tl))
                candidates = df[mask_in_tl].copy()
                if candidates.empty:
                    # затем ищем тематики, которые входят в формулировку (менее точное совпадение)
                    mask_tl_in = df["_norm"].str.contains(tl, na=False, regex=False)
                    candidates = df[mask_tl_in].copy()

                if not candidates.empty:
                    candidates["_len"] = candidates["_norm"].str.len()
                    candidates = candidates.sort_values("_len", ascending=False)

                    cat = str(candidates.iloc[0][TOPIC_MAP_COL_CAT]).strip()
                    rate = category_rate.get(cat)
                    if rate is not None:
                        notes.append(f"Тематика сопоставлена по таблице «Тематики по категориям» (вхождение): категория {cat}.")
                        return float(rate), notes



    # ---------------- 2) Эвристики (расширенные) ----------------
    def hit(*keys: str) -> bool:
        return any(k in tl for k in keys)

    # ВАЖНО: порядок правил — это качество классификации.
    # Сначала “составные” и более специфичные штуки, потом общие.

    # ---- V (музыкальная) — ловим раньше, чтобы “информационно-музыкальное” не ушло в I
    if hit("музык", "песн", "клип", "концерт", "эстрад", "популярная музыка", "музыкально", "музыка"):
        rate = category_rate.get("V", 3.0)
        notes.append("Тематика распознана эвристикой как «музыкальная» (категория V).")
        return float(rate), notes

    # ---- IV (развлекательная) — тоже довольно специфично
    if hit("развлек", "юмор", "шоу", "ток-шоу", "ток шоу", "игр", "викторин", "конкурс", "поздрав", "комед", "коморист", "розыгрыш", "молодеж"):
        rate = category_rate.get("IV", 2.7)
        notes.append("Тематика распознана эвристикой как «развлекательная» (категория IV).")
        return float(rate), notes

    # ---- II (социально-полезная/образование/дети/спорт/наука/здоровье/туризм/уроки)
    if hit(
        "дет", "для детей", "подрост", "школьник",
        "образоват", "учеб", "урок", "просветительско-образователь", "научн", "научно",
        "познавательн", "учебно-познавательн", "учебно-просветительск",
        "спорт", "спортивно", "оздоров", "здоров", "зож", "пропаганда здорового",
        "туризм"
    ):
        rate = category_rate.get("II", 2.3)
        notes.append("Тематика распознана эвристикой как «социально-полезная/образовательная/спорт/ЗОЖ» (категория II).")
        return float(rate), notes

    # ---- III (культурно-просветительская/искусство/документалистика/художественное/погода/отдых/народное)
    if hit(
        "культур", "искусств", "литератур", "поэтич", "художественн", "художественно",
        "документ", "кино", "телефильм", "передачи об искусстве",
        "народное творчество", "социально-культурн",
        "прогноз погоды", "погода", "отдых", "досуг"
    ):
        rate = category_rate.get("III", 2.5)
        notes.append("Тематика распознана эвристикой как «культурно-просветительская/художественная» (категория III).")
        return float(rate), notes

    # ---- I (информационная/новости/политика/экономика/право/религия/официальная хроника/социально-значимые)
    if hit(
        "информац", "новост", "аналит", "публицист", "общественно", "полит", "эконом", "делов",
        "правов", "социально-значим", "официальная хроника", "оперативная информация",
        "религи", "патриот", "интервью", "комментар"
    ):
        rate = category_rate.get("I", 2.0)
        notes.append("Тематика распознана эвристикой как «информационная» (категория I).")
        return float(rate), notes

    # Если вообще непонятно — дефолт
    notes.append("Тематика не распознана; применена ставка по умолчанию 2,5% (категория III).")
    return DEFAULT_TOPIC_RATE, notes


# --------------------------- loading: RKN table ---------------------------

def iter_rkn_rows(rkn_xlsx: Path) -> Tuple[List[str], Any]:
    wb = openpyxl.load_workbook(rkn_xlsx, read_only=True, data_only=True)
    ws = wb.active

    header_raw = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    while header_raw and (header_raw[-1] is None or str(header_raw[-1]).strip() == ""):
        header_raw.pop()
    header = header_raw
    max_col = len(header)

    it = ws.iter_rows(min_row=2, max_col=max_col, values_only=True)
    return header, it

@lru_cache(maxsize=1)
def _inn_to_org_map(rkn_path: str, mtime: float) -> dict:
    rkn_xlsx = Path(rkn_path)
    header, it = iter_rkn_rows(rkn_xlsx)
    idx = {h: i for i, h in enumerate(header)}

    col_inn = idx.get("ns1:inn")
    col_name = idx.get("ns1:org_name")
    if col_inn is None or col_name is None:
        return {}

    out = {}
    for row in it:
        inn = str(row[col_inn] or "").strip()
        if not inn:
            continue
        if inn not in out:
            out[inn] = str(row[col_name] or "").strip()
    return out

def get_org_name_by_inn(rkn_xlsx: Path, inn: str) -> str:
    mp = _inn_to_org_map(str(rkn_xlsx), rkn_xlsx.stat().st_mtime)
    return (mp.get(inn) or "").strip()



def build_rkn_url(license_id: str) -> str:
    return "https://rkn.gov.ru/activity/mass-media/for-broadcasters/teleradio/?id=" + quote(str(license_id), safe="")


def load_licenses_by_inn(rkn_xlsx: Path, inn: str, vars_xlsx: Path) -> Tuple[List[License], List[str]]:
    notes: List[str] = []

    header, it = iter_rkn_rows(rkn_xlsx)
    idx = {h: i for i, h in enumerate(header)}

    required = [
        "ns1:inn", "ns1:org_name", "ns1:license_num", "ns1:sreda", "ns1:population",
        "ns1:smi_name14", "ns1:smi_name", "ns1:brcst_direction", "ns1:percentage", "ns1:brcst_time"
    ]
    missing = [c for c in required if c not in idx]
    if missing:
        notes.append(f"В таблице РКН не найдены ожидаемые колонки: {missing}. Скрипт будет работать частично.")

    category_rate = build_category_rate_map(vars_xlsx)
    try:
        topics_map = pd.read_excel(vars_xlsx, sheet_name="Тематики по категориям")
        if topics_map.dropna(how="all").empty:
            topics_map = pd.DataFrame()
    except Exception:
        topics_map = pd.DataFrame()

    by_license: Dict[str, Dict[str, Any]] = {}

    def get(row, col):
        j = idx.get(col)
        if j is None:
            return None
        if j >= len(row):
            return None
        return row[j]

    for row in it:
        row_inn = str(get(row, "ns1:inn") or "").strip()
        if row_inn != inn:
            continue

        org_name = str(get(row, "ns1:org_name") or "").strip()
        lic_id = str(get(row, "ns1:license_num") or "").strip()
        sreda = str(get(row, "ns1:sreda") or "").strip()
        pop_raw = get(row, "ns1:population")

        smi14 = clean_channel_name(get(row, "ns1:smi_name14"))
        smi = clean_channel_name(get(row, "ns1:smi_name"))
        channel_name = smi14 or smi or "Неизвестный канал"

        brcst_time = get(row, "ns1:brcst_time")
        direction = str(get(row, "ns1:brcst_direction") or "").strip()
        perc = get(row, "ns1:percentage")

        if not lic_id:
            continue

        lic = by_license.setdefault(lic_id, {
            "org_name": org_name,
            "inn": inn,
            "sreda": sreda,
            "pop_values": [],
            "pop_notes": [],
            "channels": {}
        })

        pop_int, pop_notes = parse_population(pop_raw)
        if pop_int is not None:
            lic["pop_values"].append(pop_int)
        lic["pop_notes"].extend(pop_notes)

        ch = lic["channels"].setdefault(channel_name, {
            "hours": None,
            "hours_notes": [],
            "topics": []
        })

        hrs, hrs_notes = parse_hours_week(brcst_time, get(row, "ns1:smi_name"))
        if hrs is not None:
            ch["hours"] = hrs
        ch["hours_notes"].extend(hrs_notes)

        if direction:
            share = None
            if perc is not None and str(perc).strip() != "":
                try:
                    share = float(str(perc).replace(",", "."))
                except ValueError:
                    share = None

            rate, rate_notes = topic_to_rate(direction, category_rate, topics_map)
            note = "; ".join(rate_notes) if rate_notes else None
            ch["topics"].append(TopicShare(topic_raw=direction, share_pct=share, rate_pct=rate, note=note))

    licenses: List[License] = []
    for lic_id, data in by_license.items():
        media_class = normalize_media(data.get("sreda"))
        pop_total = None
        if data["pop_values"]:
            pop_total = int(sum(sorted(set(data["pop_values"]))))

        lic_obj = License(
            license_id=lic_id,
            org_name=data.get("org_name", ""),
            inn=inn,
            media_raw=data.get("sreda", ""),
            media_class=media_class,
            population_total=pop_total,
            population_notes=data.get("pop_notes", []),
            rkn_url=build_rkn_url(lic_id),
            channels=[]
        )
        for ch_name, ch_data in data["channels"].items():
            lic_obj.channels.append(Channel(
                name=ch_name,
                hours_week=ch_data.get("hours"),
                topics=ch_data.get("topics", [])
            ))
        licenses.append(lic_obj)

    if not licenses:
        notes.append("По этому ИНН в таблице РКН не найдено строк. Проверьте, что ИНН есть в выгрузке.")
    return licenses, notes


# --------------------------- computations ---------------------------

def round_rate(x: float) -> float:
    return round(x + 1e-9, 1)


def compute_license_rate(lic: License) -> Tuple[float, Dict[str, Any]]:
    det: Dict[str, Any] = {"channels": []}
    num = 0.0
    den = 0.0

    for ch in lic.channels:
        ch_rate, ch_notes = ch.avg_rate()
        hrs = ch.hours_week if ch.hours_week is not None else 168.0

        det["channels"].append({
            "channel": ch.name,
            "hours": hrs,
            "channel_rate_raw": ch_rate,
            "channel_rate": round_rate(ch_rate),
            "notes": ch_notes,
            "topics": [
                {"topic": t.topic_raw, "share_pct": t.share_pct, "rate_pct": t.rate_pct, "note": t.note}
                for t in ch.topics
            ]
        })

        num += ch_rate * hrs
        den += hrs

    if den == 0:
        return 2.5, {"warning": "Не удалось рассчитать ставку по ВЛ (нет часов/каналов). Применена 2,5%."}
    return round_rate(num / den), det


def compute_contract_rate(licenses: List[License]) -> Tuple[float, Dict[str, Any]]:
    details: Dict[str, Any] = {"licenses": []}
    num = 0.0
    den = 0.0

    for lic in licenses:
        lic_rate, lic_rate_details = compute_license_rate(lic)
        pop = lic.population_total
        hrs = lic.total_hours()
        w = (pop or 0) * hrs

        details["licenses"].append({
            "license_id": lic.license_id,
            "license_rate": lic_rate,
            "population": pop,
            "hours": hrs,
            "weight": w,
            "license_rate_details": lic_rate_details
        })

        if pop is None:
            continue
        num += lic_rate * w
        den += w

    if den == 0:
        return 2.5, {"warning": "Не удалось рассчитать взвешенную ставку (нет населения). Применена 2,5%."}
    return round_rate(num / den), details


def compute_percent_sum_q(
    contract_rate: float,
    annual_revenue: Optional[float],
    revenue_q: Optional[float],
    expenses_q: Optional[float],
) -> Tuple[Optional[float], Dict[str, Any], List[str]]:
    notes: List[str] = []
    det: Dict[str, Any] = {"base_type": None, "base_q": None}

    base_q = None
    if revenue_q is not None:
        base_q = revenue_q
        det["base_type"] = "доходы (квартал)"
        notes.append("База для процента: доходы за квартал (введено пользователем).")
    elif annual_revenue is not None:
        base_q = annual_revenue / 4.0
        det["base_type"] = "доходы (год/4)"
        notes.append("База для процента: годовая выручка/доход разделён на 4 (если нет поквартальных данных).")
    elif expenses_q is not None:
        base_q = expenses_q
        det["base_type"] = "расходы (квартал)"
        notes.append("База для процента: расходы за квартал (ветка 100% госструктура / нет доходов).")
    else:
        notes.append("Не задана база для процента (нет доходов/выручки/расходов).")
        return None, det, notes

    det["base_q"] = base_q
    return round(base_q * (contract_rate / 100.0), 2), det, notes


def lookup_min_sum(mins_df: pd.DataFrame, population: int, media_class: str) -> Optional[float]:
    sub = mins_df[
        (mins_df["Среда осуществления вещания (в эфире, по кабелю, одновременно в эфире и по кабелю)"]
         .astype(str).str.strip() == media_class)
    ].copy()
    if sub.empty:
        return None

    for _, r in sub.iterrows():
        lo = int(r["Численность населения на территории вещания, от (человек)"])
        hi = r["Численность населения на территории вещания, до (человек)"]
        hi_val = int(hi) if pd.notna(hi) else None
        if population >= lo and (hi_val is None or population <= hi_val):
            return float(r["Минимальная сумма авторского вознаграждения за квартал, рублей"])
    return None


def hour_coeff(hours_df: pd.DataFrame, hours_week: float) -> float:
    for _, r in hours_df.iterrows():
        lo = float(r["Количество часов вещания в неделю, от"])
        hi = float(r["Количество часов вещания в неделю, до"])
        if hours_week >= lo and hours_week <= hi:
            return float(r["Коэффициент к установленной минимальной сумме вознаграждения"])
    return 1.0


def discount_by_licenses(disc_df: pd.DataFrame, n_licenses: int) -> float:
    for _, r in disc_df.iterrows():
        lo = int(r["Минимальное количество вещательных лицензий одного пользователя"])
        hi = r["Максимальное количество вещательных лицензий одного пользователя"]
        hi_val = int(hi) if pd.notna(hi) else None
        if n_licenses >= lo and (hi_val is None or n_licenses <= hi_val):
            disc_pct = float(r["Размер скидки к совокупной минимальной сумме вознаграждения, процентов"])
            return 1.0 - disc_pct / 100.0
    return 1.0


def contract_period_coeff(period_df: pd.DataFrame, contract_quarter: int) -> float:
    for _, r in period_df.iterrows():
        lo = int(r["Отчетный период действия лицензионного договора, начиная с (номер квартала)"])
        hi = int(r["Отчетный период действия лицензионного договора, по (номер квартала включительно)"])
        if contract_quarter >= lo and contract_quarter <= hi:
            return float(r["Коэффициент к минимальной сумме вознаграждения в указанный период"])
    return 1.0


def compute_min_total(
    licenses: List[License],
    vars_xlsx: Path,
    annual_income_for_rules: Optional[float],
    contract_quarter: int,
    internet_resources: int,
    past_year_percent_paid: Optional[float],
    percent_sum_q: Optional[float],
    contract_media: str = "auto",
    use_small_income_branch: Optional[bool] = None,
    new_user_only: bool = False,
    assoc_member: bool = False,
    subscriber_total: Optional[int] = None,
) -> Tuple[Optional[float], Dict[str, Any], List[str]]:
    """Блок C: расчёт минималки по таблицам + поправки (население/часы/интернет/скидка по числу лицензий)
    и затем «гильотина» как автоматическое ограничение минималки по сумме по проценту.

    ВАЖНО: коэффициенты льгот/стимулов (блок E) здесь НЕ применяются.
    """
    notes: List[str] = []
    details: Dict[str, Any] = {"steps": []}

    mins_df = pd.read_excel(vars_xlsx, sheet_name="Минимальные суммы по населению")
    disc_df = pd.read_excel(vars_xlsx, sheet_name="Скидки по количеству лицензий")
    hours_df = pd.read_excel(vars_xlsx, sheet_name="Коэффициенты по часам")
    params_df = pd.read_excel(vars_xlsx, sheet_name="Параметры для расчетов")

    def get_param_contains(substr: str, default: float) -> float:
        col = "Наименование параметра для расчета авторского вознаграждения"
        if col not in params_df.columns:
            return default
        sub = params_df[params_df[col].astype(str).str.contains(substr, case=False, na=False)]
        if sub.empty:
            return default
        return float(sub.iloc[0]["Значение параметра"])

    THRESH_SMALL = get_param_contains("Порог годового дохода", 1_500_000.0)
    SMALL_K = get_param_contains("Коэффициент уменьшения", 0.5)
    SMALL_MAX_Q = int(get_param_contains("Максимальное количество отчетных периодов применения половины", 8))
    INTERNET_PCT = get_param_contains("Дополнительный процент увеличения", 0.15)
    INTERNET_MIN_ADD = get_param_contains("Минимальное увеличение", 12500.0)
    GUILLOTINE_COEF = get_param_contains("Коэффициент гильотины", 1.00)

    pops_missing = [lic.license_id for lic in licenses if lic.population_total is None]
    if pops_missing:
        notes.append(f"Не найдена численность населения по лицензиям: {pops_missing}. Без населения минималка будет неполной.")

    # определяем среду договора на уровне агрегирования
    media_classes = [lic.media_class for lic in licenses]
    if "Одновременно в эфире и по кабелю" in media_classes:
        has_two_media = True
    else:
        has_air = any("эфир" in (lic.media_raw or "").lower() or "назем" in (lic.media_raw or "").lower() for lic in licenses)
        has_cable = any("кабель" in (lic.media_raw or "").lower() for lic in licenses)
        has_two_media = bool(has_air and has_cable)

    media_for_agg = "Одновременно в эфире и по кабелю" if has_two_media else "В эфире или по кабелю"

    contract_media = (contract_media or "auto").lower().strip()
    if contract_media in ("cable", "air"):
        has_two_media = False
        media_for_agg = "В эфире или по кабелю"
        notes.append("Среда договора принудительно задана как «В эфире или по кабелю» (эфир/кабель).")
    elif contract_media == "both":
        has_two_media = True
        media_for_agg = "Одновременно в эфире и по кабелю"
        notes.append("Среда договора принудительно задана как «Одновременно в эфире и по кабелю».")

    if use_small_income_branch is not None:
        small_branch = use_small_income_branch
    else:
        small_branch = bool(
            annual_income_for_rules is not None
            and annual_income_for_rules <= THRESH_SMALL
            and contract_quarter <= SMALL_MAX_Q
        )

    min_total = 0.0

    # --- Ветка 3.4.2 / 3.5: минималка по абонентам ---
    if subscriber_total is not None:
        subs = int(subscriber_total)
        if subs < 0:
            return None, details, notes + ["Количество абонентов не может быть отрицательным."]
    
        min_total = float(subs) * 5.0
        details["steps"].append({"step": "SUBSCRIBERS_MIN", "subscriber_total": subs, "min_after": min_total})
        notes.append("Минимальная сумма рассчитана по абонентам: не менее 5 руб. за абонента (пп. 3.4.2 / 3.5).")
    
        # интернет-доплата (3.6) — оставляем как у тебя ниже, она применится дальше
    else:
        min_total = 0.0

    if subscriber_total is None:
    # C1/C2/C7
        if small_branch:
            N_sum = sum(int(lic.population_total) for lic in licenses if lic.population_total is not None)
            if N_sum <= 0:
                return None, details, notes + ["Нельзя применить ветку малого дохода: нет суммарной численности населения."]
            m = lookup_min_sum(mins_df, N_sum, media_for_agg)
            if m is None:
                return None, details, notes + ["Не найдена минималка в таблице по суммарной численности населения."]
            min_total = SMALL_K * m
            details["steps"].append({"step": "C3", "N_sum": N_sum, "media": media_for_agg, "min_table": m, "k_small": SMALL_K, "min_after": min_total})
            notes.append("Включена ветка малого дохода: минималка по суммарной численности населения и затем ×0,5.")
        else:
            per_lic = []
            for lic in licenses:
                if lic.population_total is None:
                    continue
    
                media_for_min = lic.media_class
                if contract_media in ("cable", "air"):
                    media_for_min = "В эфире или по кабелю"
                elif contract_media == "both":
                    media_for_min = "Одновременно в эфире и по кабелю"
    
                m = lookup_min_sum(mins_df, int(lic.population_total), media_for_min)
                if m is None:
                    notes.append(f"Не найдена минималка по таблице для лицензии {lic.license_id} (население={lic.population_total}, среда={media_for_min}).")
                    continue
    
                hrs = lic.total_hours()
                coeff = 1.0
                if hrs < 126:
                    coeff = hour_coeff(hours_df, hrs)
    
                m2 = m * coeff
                per_lic.append({"license_id": lic.license_id, "population": lic.population_total, "media": media_for_min, "min_table": m, "hours_week": hrs, "hour_coeff": coeff, "min_after": m2})
                min_total += m2
    
            details["steps"].append({"step": "C1+C2(+C7)", "per_license": per_lic, "min_after": min_total})

    # C4 скидка по числу лицензий
    n_lic = len(licenses)
    if n_lic > 3:
        k = discount_by_licenses(disc_df, n_lic)
        min_total *= k
        details["steps"].append({"step": "C4", "n_licenses": n_lic, "coeff": k, "min_after": min_total})
        notes.append(f"Применена скидка по числу лицензий (кол-во ВЛ={n_lic}).")

    # C6 интернет-доплата
    if internet_resources and internet_resources > 0:
        add_per = max(INTERNET_PCT * min_total, INTERNET_MIN_ADD)
        delta = add_per * internet_resources
        min_total += delta
        details["steps"].append({"step": "C6", "resources": internet_resources, "add_per_resource": add_per, "delta": delta, "min_after": min_total})
        notes.append("Добавлена доплата за интернет-вещание (+15%, но не менее 12 500 за ресурс).")

    # «Гильотина» — НЕ отдельный режим. Это автоматический срез минималки по сумме по проценту.
    # Логируем шаг всегда (если известна сумма по %), даже если среза не произошло.
    if percent_sum_q is not None:
        min_before = float(min_total)
        cap = float(GUILLOTINE_COEF) * float(percent_sum_q)
        min_after = min(min_before, cap)

        # применяем срез
        min_total = min_after

        details["steps"].append({
            "step": "GUILLOTINE",
            "min_before": round(min_before, 2),
            "s_percent": round(float(percent_sum_q), 2),
            "coef": float(GUILLOTINE_COEF),
            "cap": round(cap, 2),
            "min_after": round(min_after, 2),
        })

        if abs(min_after - min_before) > 1e-9:
            notes.append("Гильотина: ограничение минималки по % (min = min(min, coef*s_percent)).")

    return round(min_total, 2), details, notes


# --------------------------- reporting ---------------------------

def money(x: Optional[float]) -> str:
    if x is None:
        return "—"
    return f"{x:,.2f}".replace(",", " ").replace(".00", "")


def format_report(
    inn: str,
    year: Optional[int],
    annual_revenue: Optional[float],
    revenue_q: Optional[float],
    expenses_q: Optional[float],
    internet_resources: int,
    contract_quarter: int,
    new_user_only: bool,
    assoc_member: bool,
    licenses: List[License],
    contract_rate: float,
    percent_sum_q: Optional[float],
    min_total: Optional[float],
    notes: List[str],
    needs: List[str],
) -> str:
    lines: List[str] = []

    org_name = licenses[0].org_name if licenses else "Не найдено (нет записей в РКН)"
    lines.append(f"{org_name}, ИНН {inn}.")
    lines.append("")

    if revenue_q is not None:
        lines.append(f"Доходы за квартал (введено): {money(revenue_q)} ₽.")
    elif annual_revenue is not None:
        lines.append(f"Выручка/доход за {year or 'год'} (введено): {money(annual_revenue)} ₽.")
        lines.append("База квартала: год/4 (если нет поквартальных данных).")
    elif expenses_q is not None:
        lines.append(f"Расходы за квартал (введено): {money(expenses_q)} ₽ (ветка 100% госструктура / нет доходов).")
    else:
        lines.append("Финансовая база для расчёта процента: НЕ ЗАДАНА.")
    lines.append("")

    lines.append(f"Ранее работал с РАО: {'да' if (not new_user_only) else 'нет'}.")
    lines.append(f"Член отраслевой ассоциации: {'да' if assoc_member else 'нет'}.")
    lines.append("")

    lines.append(f"Вещательные лицензии (по таблице РКН): {len(licenses)} шт.")
    for lic in licenses:
        pop = lic.population_total
        pop_str = f"{pop:,}".replace(",", " ") if pop is not None else "не найдено"
        lines.append(f"— Номер лицензии (РКН): {lic.license_id}; среда: {lic.media_raw} → {lic.media_class}; население: {pop_str}.")
        lines.append(f"  РКН: {lic.rkn_url}")
        if lic.population_notes:
            for n in lic.population_notes[:2]:
                lines.append(f"  Примечание: {n}")

        for ch in lic.channels[:10]:
            hrs = ch.hours_week if ch.hours_week is not None else 168.0
            ch_rate, ch_notes = ch.avg_rate()
            lines.append(f"  • Канал/СМИ: {ch.name}; часы/нед: {hrs:g}; ставка канала: {round_rate(ch_rate):.1f}%.")
            for tn in ch_notes[:2]:
                lines.append(f"    — {tn}")
            if ch.topics:
                for t in ch.topics[:8]:
                    share = f"{t.share_pct:g}%" if t.share_pct is not None else "без доли"
                    lines.append(f"    Тематика: {t.topic_raw} ({share}) → {t.rate_pct:.1f}%.")

    lines.append("")
    lines.append(f"Процентная ставка по договору (взвешенная по часам×населению): {contract_rate:.1f}%.")
    lines.append("")
    lines.append(f"Расчётная сумма по проценту за квартал: {money(percent_sum_q)} ₽.")
    lines.append("")
    lines.append(f"Минимальная сумма за квартал (по правилам/таблицам): {money(min_total)} ₽.")
    lines.append("")

    if percent_sum_q is not None and min_total is not None:
        pay = max(percent_sum_q, min_total)
        which = "по проценту" if percent_sum_q >= min_total else "по минималке"
        lines.append(f"Итог: {contract_rate:.1f}% от базы за квартал, но не менее {money(min_total)} ₽. К оплате: {money(pay)} ₽ ({which}).")
    else:
        lines.append("Итог: недостаточно данных для финального вывода (см. «Нужно уточнить»).")
    lines.append("")

    if internet_resources:
        lines.append(f"Интернет-вещание: указано ресурсов — {internet_resources} (применена доплата по правилам).")
        lines.append("")

    if needs:
        lines.append("Нужно уточнить/проверить:")
        for x in needs:
            lines.append(f"— {x}")
        lines.append("")

    if notes:
        lines.append("Примечания и допущения:")
        for n in notes:
            lines.append(f"— {n}")
        lines.append("")

    return "\n".join(lines)


# --------------------------- main (non-interactive) ---------------------------

def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--inn", required=True, help="ИНН (10/12 цифр)")
    ap.add_argument("--year", type=int, default=None, help="Год (для подписи в отчёте)")

    ap.add_argument("--annual_revenue", type=float, default=None, help="Годовая выручка/доход")
    ap.add_argument("--revenue_q", type=float, default=None, help="Доходы за квартал")
    ap.add_argument("--expenses_q", type=float, default=None, help="Расходы за квартал (ветка госструктуры)")

    ap.add_argument("--internet_resources", type=int, default=0)
    ap.add_argument("--contract_quarter", type=int, default=1)
    ap.add_argument("--contract_media", type=str, default="auto", choices=["auto", "cable", "air", "both"])

    ap.add_argument("--new_user", action="store_true")
    ap.add_argument("--assoc_member", action="store_true")

    ap.add_argument("--only_license", type=str, default=None)
    ap.add_argument("--licenses", action="append", default=None, help="Номера лицензий (можно несколько, через запятую)")

    ap.add_argument("--rkn_xlsx", type=str, default="Таблица РКН.xlsx")
    ap.add_argument("--vars_xlsx", type=str, default="Переменные из ставок.xlsx")

    ap.add_argument("--force_small_income", action="store_true")
    ap.add_argument("--no_small_income", action="store_true")

    ap.add_argument("--population_override", type=int, default=None)
    ap.add_argument("--population_by_license", action="append", default=None, help="Переопределение населения: ЛИЦЕНЗИЯ=НАСЕЛЕНИЕ")
    ap.add_argument("--subscriber_total", type=int, default=None, help="Суммарное количество абонентов (ветка 3.4.2/3.5)")

    ap.add_argument("--non_interactive", action="store_true", help="для сайта всегда ставим этот флаг")

    args = ap.parse_args(argv)

    p = Progress(enabled=not bool(args.non_interactive))

    try:
        inn = parse_inn(args.inn)
    except Exception as e:
        print(f"Ошибка: {e}")
        return 2

    base_dir = Path(__file__).resolve().parent
    rkn_xlsx = Path(args.rkn_xlsx)
    vars_xlsx = Path(args.vars_xlsx)
    if not rkn_xlsx.is_absolute():
        rkn_xlsx = base_dir / rkn_xlsx
    if not vars_xlsx.is_absolute():
        vars_xlsx = base_dir / vars_xlsx

    if not rkn_xlsx.exists():
        for name in (
            "Таблица РКН slim.xlsx",
            "Таблица РКН очищенная.xlsx",
            "Таблица РКН (2).xlsx",
        ):
            alt = rkn_xlsx.with_name(name)
            if alt.exists():
                rkn_xlsx = alt
                break

    if not rkn_xlsx.exists():
        print(f"Ошибка: не найден файл РКН: {rkn_xlsx}")
        return 2
    if not vars_xlsx.exists():
        print(f"Ошибка: не найден файл ставок: {vars_xlsx}")
        return 2

    p.tick("читаю РКН и собираю лицензии")
    licenses, load_notes = load_licenses_by_inn(rkn_xlsx, inn, vars_xlsx)

    selected_ids = parse_license_list(args.licenses, args.only_license)
    if selected_ids:
        selected_set = {str(x).strip() for x in selected_ids}
        licenses = [x for x in licenses if str(x.license_id).strip() in selected_set]
        if not licenses:
            print(f"Ошибка: не найдены лицензии {sorted(selected_set)} у этого ИНН в таблице РКН.")
            return 2

    pop_by_license = parse_population_by_license(args.population_by_license)

    if args.population_override is not None:
        po = int(args.population_override)
        for lic in licenses:
            if str(lic.license_id).strip() in pop_by_license:
                continue
            old = lic.population_total
            lic.population_total = po
            note = f"Переопределено пользователем: {po}" + (f" (РКН: {old})" if old is not None else "")
            lic.population_notes.append(note)

    if pop_by_license:
        for lic in licenses:
            key = str(lic.license_id).strip()
            if key not in pop_by_license:
                continue
            po = int(pop_by_license[key])
            old = lic.population_total
            lic.population_total = po
            note = f"Переопределено пользователем по лицензии: {po}" + (f" (РКН: {old})" if old is not None else "")
            lic.population_notes.append(note)

    needs: List[str] = []
    notes: List[str] = []
    notes.extend(load_notes)

    if not licenses:
        print("Ошибка: нет данных по ИНН в таблице РКН.")
        return 2

    if all(lic.population_total is None for lic in licenses):
        needs.append("В РКН-таблице не заполнено население. Нужно взять численность населения территории вещания из карточек РКН.")

    p.tick("считаю процентную ставку по договору")
    contract_rate, _ = compute_contract_rate(licenses)

    p.tick("считаю сумму по проценту за квартал")
    percent_sum_q, _, percent_notes = compute_percent_sum_q(
        contract_rate=contract_rate,
        annual_revenue=args.annual_revenue,
        revenue_q=args.revenue_q,
        expenses_q=args.expenses_q,
    )
    notes.extend(percent_notes)

    if percent_sum_q is None:
        needs.append("Нужна финансовая база: годовая выручка/доход или доходы за квартал или расходы за квартал (для ветки госструктуры).")

    annual_income_for_rules = None
    if args.annual_revenue is not None:
        annual_income_for_rules = float(args.annual_revenue)
    elif args.revenue_q is not None:
        annual_income_for_rules = float(args.revenue_q) * 4.0

    if args.force_small_income and args.no_small_income:
        print("Ошибка: нельзя одновременно --force_small_income и --no_small_income")
        return 2

    use_small_income = None
    if args.force_small_income:
        use_small_income = True
    elif args.no_small_income:
        use_small_income = False

    p.tick("считаю минимальную сумму")
    p.tick("считаю минимальную сумму за квартал")
    min_total, min_details, min_notes = compute_min_total(
        licenses=licenses,
        vars_xlsx=vars_xlsx,
        annual_income_for_rules=annual_income_for_rules,
        contract_quarter=args.contract_quarter,
        internet_resources=args.internet_resources,
        past_year_percent_paid=None,
        percent_sum_q=percent_sum_q,
        contract_media=args.contract_media,
        use_small_income_branch=use_small_income,
        new_user_only=bool(args.new_user),
        assoc_member=bool(args.assoc_member),
        subscriber_total=args.subscriber_total,

    )
    notes.extend(min_notes)

    # Ветка малого дохода используется только если явно включена логикой блока C.
    small_branch_used = any("Включена ветка малого дохода" in n for n in notes)

    # Блок E: льготы/стимулы применяем И к минималке, И к сумме по проценту
    # (чтобы не было ситуации, когда % уменьшился, а минималка осталась без изменений — и наоборот).
    try:
        down_df = pd.read_excel(vars_xlsx, sheet_name="Понижающие коэффициенты")
    except Exception:
        down_df = None

    def down_coeff(contains: str, default: float = 1.0) -> float:
        if down_df is None:
            return default
        col = "Условие применения понижающего коэффициента (описательное обозначение)"
        if col not in down_df.columns:
            return default
        sub = down_df[down_df[col].astype(str).str.contains(contains, case=False, na=False)]
        if sub.empty:
            return default
        return float(sub.iloc[0]["Понижающий коэффициент к минимальной сумме вознаграждения"])

    cfE = 1.0
    # По примерам: понижающие коэффициенты «новый пользователь/период договора» применяем только в ветке малого дохода.
    if bool(args.new_user) and small_branch_used:
        # коэффициент «новый пользователь»
        # (по таблице обычно разные формулировки для «одновременно» и «эфир/кабель»)
        media_classes = [lic.media_class for lic in licenses]
        has_two_media = ("Одновременно в эфире и по кабелю" in media_classes)
        if not has_two_media:
            has_air = any("эфир" in (lic.media_raw or "").lower() or "назем" in (lic.media_raw or "").lower() for lic in licenses)
            has_cable = any("кабель" in (lic.media_raw or "").lower() for lic in licenses)
            has_two_media = bool(has_air and has_cable)
        media_for_agg = "Одновременно в эфире и по кабелю" if has_two_media else "В эфире или по кабелю"

        if args.contract_media in ("cable", "air"):
            media_for_agg = "В эфире или по кабелю"
        elif args.contract_media == "both":
            media_for_agg = "Одновременно в эфире и по кабелю"

        if media_for_agg == "Одновременно в эфире и по кабелю":
            cfE *= down_coeff("одновременно", 1.0)
        else:
            cfE *= down_coeff("Новый пользователь, заключающий", 1.0)

        # коэффициент по периоду действия договора (если таблица/лист есть)
        period_df = None
        for sheet in (
            "Коэффициенты по периодам",
            "Коэффициенты по периодам действия договора",
            "Коэффициенты по периоду действия договора",
            "Коэффициенты по периодам действия лицензионного договора",
        ):
            try:
                period_df = pd.read_excel(vars_xlsx, sheet_name=sheet)
                break
            except Exception:
                period_df = None

        if period_df is not None:
            try:
                cf_period = contract_period_coeff(period_df, int(args.contract_quarter))
                cfE *= float(cf_period)
            except Exception:
                pass

    if bool(args.assoc_member):

        cfE *= down_coeff("ассоциаци", 1.0)

    if cfE != 1.0:
        if min_total is not None:
            min_total = round(float(min_total) * cfE, 2)
        if percent_sum_q is not None:
            percent_sum_q = round(float(percent_sum_q) * cfE, 2)
        notes.append(f"Блок E: применён общий коэффициент льгот/стимулов cfE={cfE:g} к % и к минималке.")

    p.tick("формирую отчёт")
    report = format_report(
        inn=inn,
        year=args.year,
        annual_revenue=args.annual_revenue,
        revenue_q=args.revenue_q,
        expenses_q=args.expenses_q,
        internet_resources=args.internet_resources,
        contract_quarter=args.contract_quarter,
        new_user_only=bool(args.new_user),
        assoc_member=bool(args.assoc_member),
        licenses=licenses,
        contract_rate=contract_rate,
        percent_sum_q=percent_sum_q,
        min_total=min_total,
        notes=notes,
        needs=needs,
    )
    print(report)
    return 0


def run_calc_capture(argv: List[str]) -> Tuple[int, str]:
    """
    Запускает main(argv=...), возвращает (exit_code, stdout_text).
    """
    buf = io.StringIO()
    try:
        with redirect_stdout(buf):
            code = int(main(argv))
    except SystemExit as e:
        code = int(getattr(e, "code", 1) or 0)
    except Exception as e:
        code = 2
        buf.write(f"Ошибка: {type(e).__name__}: {e}\n")
    return code, buf.getvalue()


if __name__ == "__main__":
    raise SystemExit(main())

def fix_mojibake(s: str) -> str:
    """
    Чинит типичный случай: UTF-8 байты были интерпретированы как latin-1.
    Если строка нормальная — вернёт как есть.
    """
    if not s:
        return s
    if "Ð" not in s and "Ñ" not in s:
        return s
    try:
        return s.encode("latin1").decode("utf-8")
    except Exception:
        return s
